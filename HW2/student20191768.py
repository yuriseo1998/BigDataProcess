#!/usr/bin/python3
import openpyxl
import pandas as pd
wb = openpyxl.load_workbook( "student.xlsx" )
ws = wb['Sheet1']
row_id = 1
for row in ws:
	if row_id != 1:
		sum_v = ws.cell(row = row_id, column = 3).value * 0.3
		sum_v += ws.cell(row = row_id, column = 4).value * 0.35
		sum_v += ws.cell(row = row_id, column = 5).value * 0.34
		sum_v += ws.cell(row = row_id, column = 6).value
		ws.cell(row = row_id, column = 7).value = sum_v
	row_id += 1
wb.save( "student.xlsx" )
# grade
df = pd.read_excel("student.xlsx", engine = "openpyxl", names=['id', "name", "midterm","final","homework","attendance","total","grade"], sheet_name = "Sheet1")
df = df.sort_values(by="total", ascending=False)
df = df.reset_index(drop=True)
ap = df.loc[round(len(df) * 0.15),'total']
a = df.loc[round(len(df) * 0.3),'total']
bp = df.loc[round(len(df) * 0.5),'total']
b = df.loc[round(len(df) * 0.7),'total']
cp = df.loc[int(len(df) * 0.85),'total']
wb = openpyxl.load_workbook( "student.xlsx" )
ws = wb['Sheet1']
row_id = 1
for row in ws:
	if row_id != 1:
		score = ws.cell(row = row_id, column = 7).value
		if score <= cp:
			ws.cell(row = row_id, column = 8).value = 'C0'
		elif score <= b:
			ws.cell(row = row_id, column = 8).value = 'C+'
		elif score <= bp:
			ws.cell(row = row_id, column = 8).value = 'B0'
		elif score <= a:
			ws.cell(row = row_id, column = 8).value = 'B+'
		elif score <= ap:
			ws.cell(row = row_id, column = 8).value = 'A0'
		else:
			ws.cell(row = row_id, column = 8).value = 'A+'
	row_id += 1
wb.save( "student.xlsx" )
